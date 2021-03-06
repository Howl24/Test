import csv
import codecs
import re,string
import math
import unicodedata
import uuid
import openpyxl
from cassandra.cluster import Cluster
from nltk.corpus import stopwords
from nltk.stem.snowball import SpanishStemmer

class generadorReportes():

	areas_funcionales=None
	reportes=None
	especialidades=None
	columnas={}
	sheetAvisos=None	

	def __init__(self):
		cluster=Cluster()
		session = cluster.connect('btpucp')
		self.areas_funcionales=session.execute("select * from areas_funcionales")
		self.reportes=session.execute("select * from reportes")
		#self.especialidades=session.execute("select * from especialidades")


	def leerExcel(self,nombArch):
		wb=openpyxl.load_workbook(nombArch)
		sheets=wb.get_sheet_names()
		self.sheetAvisos=wb.get_sheet_by_name(sheets[0])
		maxColumnas=self.sheetAvisos.max_column+1
		#print(maxColumnas)
		for i in range(1,maxColumnas):
			self.columnas[self.sheetAvisos.cell(row=1,column=i).value]=i
		#print(self.columnas)

	def _find_word(self,text,search):
		textSplited=text.split()
		searchSplited = search.split()

		cnt = 0
		for search_word in searchSplited:
			if search_word in textSplited:
				cnt+=1
	
		if (cnt==len(searchSplited)):
			return True
		else:	
			return False
			
		
	def _remove_punctuacion(self,text):
		regex = re.compile('[%s]' % re.escape(string.punctuation))
		return regex.sub(' ', text) 

	def _remove_numbers(self,text):
		return ''.join([letter for letter in text if not letter.isdigit()])

	def _remove_accents(self,data):
		return ''.join(x for x in unicodedata.normalize('NFKD', data) if x in string.ascii_letters).lower()

	def preprocesamiento(self,text):
		stopEnglish = stopwords.words('english')
		stopSpanish = stopwords.words('spanish')
		stopSpanish.append("y/o")
		stopSpanish.append("–")
		stopSpanish.append("•")
		stm=SpanishStemmer() 

		text=text.lower()
		text=self._remove_punctuacion(text)
		text=self._remove_numbers(text) 
		
		textProcess=''
		firstWord=True
		for word in text.split():			
			if word not in stopEnglish and word not in stopSpanish:
				#word=self.remove_accents(word)
				word=stm.stem(word)
				if firstWord:
					textProcess+=word
					firstWord=False
				else:
					textProcess+=' '+word
		return textProcess

	def reporteAreasFuncionales(self):
		conteoSubareas={}
		stemSubareas={}
		for area in self.areas_funcionales:
			similares={}
			if area.similares!=None:
				similares=area.similares
			for subarea in area.subareas:
				conteoSubareas[subarea]=0
				stemSubareas[subarea]=[self.preprocesamiento(subarea)]
				if subarea in similares.keys():
					for similar in similares[subarea]:
						stemSubareas[subarea].append(self.preprocesamiento(similar))
		
		
		listaColumnas=[]
		listaColumnas.append(self.columnas['Job: Job Title'])
		#listaColumnas.append(self.columnas['Job_Description'])
		#listaColumnas.append(self.columnas['Job_Qualifications'])
		
		maxFilas=self.sheetAvisos.max_row+1
		numOfertasRepetidas=0
		for numOferta in range(2,maxFilas):
			text=''
			for numColumn in listaColumnas:
				text+=str(self.sheetAvisos.cell(row=numOferta,column=numColumn).value)
				text+='\n'		
			text=self.preprocesamiento(text)
			#print(self.sheetAvisos.cell(row=numOferta,column=1).value,text)
			#break

			firstInOffer=True
			firstInSubArea=False
			for subarea in sorted(conteoSubareas.keys()):
				for stemWord in stemSubareas[subarea]:
					if text.find(stemWord)!=-1:
						if firstInSubArea and firstInOffer:
							numOfertasRepetidas+=1
							#print(self.sheetAvisos.cell(row=numOferta,column=1).value,text)
							firstInOffer=False
						conteoSubareas[subarea]+=1
						firstInSubArea=True
						break
			
		return conteoSubareas

		#print(conteoSubareas)
		print("Numero de ofertas con mas de 1 area funcional",numOfertasRepetidas)
		for key in sorted(conteoSubareas.keys()):
			print("%s: %d"%(key,conteoSubareas[key]))

	def reporte_Columnas(self,nombReporte):
		listaColumnas=[]
		if nombReporte=='Idiomas':
			listaColumnas.append('Job: Language')
			listaColumnas.append('Job: Qualifications')
		elif nombReporte=='Caracteristicas':
			listaColumnas.append('Job: Description')
			listaColumnas.append('Job: Qualifications')
		elif nombReporte=='Estudios':
			listaColumnas.append('Job: Degree Level')
			listaColumnas.append('Job: Qualifications')
		elif nombReporte=='Responsabilidades':
			listaColumnas.append('Job: Description')
			listaColumnas.append('Job: Qualifications')
		elif nombReporte=='Cargos':
			listaColumnas.append('Job: Position Level')
			#listaColumnas.append('Job: Description')
			#listaColumnas.append('Job: Qualifications')
		elif nombReporte=='Tam_empresa':
			listaColumnas.append('Dimension empresas')
		elif nombReporte=='Competencias':
			listaColumnas.append('Job: Description')
			listaColumnas.append('Job: Qualifications')
		elif nombReporte=='Software':
			listaColumnas.append('Job: Software')
			listaColumnas.append('Job: Qualifications')

		return listaColumnas

	def reporte_Reportes(self,nombReporte):

		reporteEncontrado=None
		for reporte in self.reportes:
			if reporte.nombre==nombReporte:
				reporteEncontrado=reporte
				break

		conteoPalabras={}
		stemPalabras={}
		similares={}
		if reporteEncontrado.similares!=None:
			similares=reporteEncontrado.similares #si se usa el diccionario directamente del objeto extraido de la consulta en cassandra, se cae
		for palabra in reporteEncontrado.palabras:
			conteoPalabras[palabra]=0
			stemPalabras[palabra]=[self.preprocesamiento(palabra)] #se crea la lista de sinonimos ya 'stemmeados'
			if palabra in similares.keys(): #si existe similares de esa palabra, se agregan a la lista
				for similar in similares[palabra]:
					stemPalabras[palabra].append(self.preprocesamiento(similar))

		print(conteoPalabras)
		print(stemPalabras)
		columnasAbuscar=self.reporte_Columnas(nombReporte) #se buscan las columnas correspondientes segun el tipo de reporte

		#Revisar si hay esas columnas en los diccionarios
		listaColumnas=[]
		for nombColumnas in columnasAbuscar:
			listaColumnas.append(self.columnas[nombColumnas])

		maxFilas=self.sheetAvisos.max_row+1
		for numOferta in range(2,maxFilas):
			text=''
			for numColumn in listaColumnas:
				text+=str(self.sheetAvisos.cell(row=numOferta,column=numColumn).value)
				text+=' '
			#print(numOferta,text,"ANTES DE STEM")			
			text=self.preprocesamiento(text)
			#print(numOferta,text,"DESPUES DE STEM")

			for palabra in sorted(conteoPalabras.keys()):
				for stemWord in stemPalabras[palabra]:
					if self._find_word(text,stemWord):
						print(text,stemWord)
						conteoPalabras[palabra]+=1
						#print(stemWord,"STEMWORD")
						break

		return conteoPalabras

		

	def reporte_Sectores_Economicos(self):

		colRUC=self.columnas['RUC']
		colCIIU1=self.columnas['Descripcion1']
		colCIIU4=self.columnas['Descripcion4']

		sectorXaviso={}
		sectorXempresa={}

		maxFilas=self.sheetAvisos.max_row+1
		for numOferta in range(2,maxFilas):
			if str(self.sheetAvisos.cell(row=numOferta,column=colCIIU4).value)!='Consultora De RRHH':
				sectorEconomico=str(self.sheetAvisos.cell(row=numOferta,column=colCIIU1).value)

				if sectorEconomico in sectorXaviso.keys() and sectorEconomico!='=#N/A':
					sectorXaviso[sectorEconomico]+=1
				else:
					sectorXaviso[sectorEconomico]=1
				RUC=str(self.sheetAvisos.cell(row=numOferta,column=colRUC).value)
				if sectorEconomico in sectorXempresa.keys():
					if RUC not in sectorXempresa[sectorEconomico]:
						sectorXempresa[sectorEconomico].append(RUC)
				else:
					sectorXempresa[sectorEconomico]=[RUC]


		totalAvisos=sum(sectorXaviso.values())
		totalEmpresas=0
		for empresas in sectorXempresa.values():
			totalEmpresas+=len(empresas)
		
		for sector in sorted(sectorXaviso.keys()):
			print("Sector Economico por avisos - %s: %d - %.2f"%(sector,sectorXaviso[sector],sectorXaviso[sector]/totalAvisos*100))
			print("Sector Economico por empresas - %s: %d - %.2f"%(sector,len(sectorXempresa[sector]),len(sectorXempresa[sector])/totalEmpresas*100))
		print("Total Sectores",totalAvisos)
		print("Total Empresas",totalEmpresas)

	def hacerReportes(self,nombArchivo):
		listaReportes=['Idiomas','Caracteristicas','Estudios','Responsabilidades','Cargos','Softwares']
		#self.reporte_Sectores_Economicos()

		for reporte in listaReportes:
			conteoPalabras=self.reporte_Reportes(reporte)
			self.escribir_Excel_Reportes(nombArchivo,reporte,conteoPalabras)

		#conteo=self.reporteAreasFuncionales()
		#print(conteo)
		#self.escribir_Excel_areasFuncionales(nombArchivo,conteo)

		#conteoPalabras=self.reporte_Reportes("Caracteristicas")
		#conteoPalabras=self.reporte_Reportes("Estudios")
		#conteoPalabras=self.reporte_Reportes("Responsabilidades")
		#conteoPalabras=self.reporte_Reportes("Cargos")
		#conteoPalabras=self.reporte_Reportes("Tam_empresa")
		#conteoPalabras=self.reporte_Reportes("Competencias")
		#conteoPalabras=self.reporte_Reportes("Software")
		

	def escribir_Excel_Reportes(self,nombArchivo,tipoReporte,conteoPalabras):
		wb=openpyxl.Workbook()
		sheet=wb.active
		sheet['A1']=tipoReporte
		sheet['B1']='Cantidad'
		actualRow=2
		conteoTotal=0
		for palabra in sorted(conteoPalabras.keys()):
			sheet.cell(row=actualRow,column=1).value=palabra
			sheet.cell(row=actualRow,column=2).value=conteoPalabras[palabra]
			conteoTotal+=conteoPalabras[palabra]
			actualRow+=1
		sheet.cell(row=actualRow,column=1).value='Total'
		sheet.cell(row=actualRow,column=2).value=conteoTotal

			#print("%s: %d"%(palabra,conteoPalabras[palabra]))

		wb.save(nombArchivo+'_'+tipoReporte+'_Reporte.xlsx')

	def escribir_Excel_areasFuncionales(self,nombArchivo,conteoSubareas):
		wb=openpyxl.Workbook()
		sheet=wb.active
		sheet['A1']='Area'
		sheet['B1']='Subarea'
		sheet['C1']='Conteo'
		actualRow=2
		colArea=1
		colSubarea=2
		colConteo=3

		for area in self.areas_funcionales:
			nombArea=area.nombre
			conteoArea=0
			for subarea in area.subareas:
				sheet.cell(row=actualRow,column=colArea).value=nombArea
				sheet.cell(row=actualRow,column=colSubarea).value=subarea
				sheet.cell(row=actualRow,column=colConteo).value=conteoSubareas[subarea]
				conteoArea+=conteoSubareas[subarea]
				actualRow+=1
			sheet.cell(row=actualRow,column=colArea).value=nombArea
			sheet.cell(row=actualRow,column=colSubarea).value='Total'
			sheet.cell(row=actualRow,column=colConteo).value=conteoArea


		wb.save(nombArchivo+'_Areas_Funcionales'+'.xlsx')

		
		
generador=generadorReportes()
nombArchivo='Economia - 2015'
generador.leerExcel(nombArchivo+'.xlsx')
generador.hacerReportes(nombArchivo)


